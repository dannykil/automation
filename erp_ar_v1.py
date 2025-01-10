from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.chrome.options import Options

from webdriver_manager.core import driver

import logger
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from dto import AR
import json
import os
import natsort # 파일명 정렬
# from install import check_install_YN 
import importlib as imp
import imp

def processing_ERP_AR_Transaction(count, issue_cnt, not_issue_cnt, re_count):

    # if count == 0:
    if re_count == 1:
        logger.LoggerFactory.create_logger()

    with open('./conf/config.json', 'r') as f:
        install_config = json.load(f)
    
    print(install_config["installed"])
    if install_config["installed"] == "Y":    

        try:
            print('1st try')
            depth1 = '1) 설치여부 확인 및 사용자 정보 가져오기'
            depth2 = '2) 엑셀에 등록된 데이터 기반 json 데이터 생성'
            depth3 = '3) ERP 로그인 및 Receivables > Billing 이동'
            depth4 = '4) Transaction 생성 및 General Information 등록'
            depth5 = '5) Invoice Lines 및 Distribution 등록'

            logger.LoggerFactory._LOGGER.info('ERP AR 발행 Process 및 현재 위치 - 5 depth')
            logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
            logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
            logger.LoggerFactory._LOGGER.info('1) 설치여부 확인 및 사용자 정보 가져오기 <<< [현재위치]')
            logger.LoggerFactory._LOGGER.info('2) 엑셀에 등록된 데이터 기반 json 데이터 생성')
            logger.LoggerFactory._LOGGER.info('3) ERP 로그인 및 Receivables > Billing 이동')
            logger.LoggerFactory._LOGGER.info('4) Transaction 생성 및 General Information 등록')
            logger.LoggerFactory._LOGGER.info('5) Invoice Lines 및 Distribution 등록')

            with open('./conf/config.json') as f:
                user_config = json.load(f)

            installed   = user_config['installed']

            # 1) 설치여부 확인 및 사용자 정보 가져오기
            if installed == 'Y':

                user        = user_config['user']
                account     = user_config['account']
                password    = user_config['password']
                destination = user_config['filepath'] + user_config['filename']
                displayYN   = user_config['displayYN']

                logger.LoggerFactory._LOGGER.info('user        = {}'.format(user))
                logger.LoggerFactory._LOGGER.info('account     = {}'.format(account))
                logger.LoggerFactory._LOGGER.info('password    = {}'.format(password))
                logger.LoggerFactory._LOGGER.info('installed   = {}'.format(installed))
                logger.LoggerFactory._LOGGER.info('destination = {}'.format(destination))

                backup = './backup/' + datetime.now().strftime('%Y') + '/' + datetime.now().strftime('%m') + '/' + datetime.now().strftime('%d') + '/' + datetime.now().strftime('%H%M') + '/'

            else:
                raise Exception

        except Exception as e:
            print('Exception')
            logger.LoggerFactory._LOGGER.info('error msg : {}'.format(e))
            logger.LoggerFactory._LOGGER.info('1) 설치여부 확인 및 사용자 정보 가져오기 <<< 실패')
            logger.LoggerFactory._LOGGER.info('설치를 통해 사용자 정보 등록 부탁드립니다.')
            re_count = re_count + 1
        
        else:
            print('1st else')

            # try-except-else 내 다시 try문 사용가능 x 2
            try:
                print('else - 2nd try')
                logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                logger.LoggerFactory._LOGGER.info('1) 설치여부 확인 및 사용자 정보 가져오기')
                logger.LoggerFactory._LOGGER.info('2) 엑셀에 등록된 데이터 기반 json 데이터 생성 <<< [현재위치]')
                logger.LoggerFactory._LOGGER.info('3) ERP 로그인 및 Receivables > Billing 이동')
                logger.LoggerFactory._LOGGER.info('4) Transaction 생성 및 General Information 등록')
                logger.LoggerFactory._LOGGER.info('5) Invoice Lines 및 Distribution 등록')

                # with open('./conf/template_GeneralInformation.json') as template_GeneralInformation:
                #     json_GI = json.load(template_GeneralInformation)

                # jsonObject가 중복됨
                # with open('./conf/template_InvoiceDetails.json') as template_InvoiceDetails:
                #     json_ID = json.load(template_InvoiceDetails)

                # ★AR종합 가져오기
                wb = load_workbook(destination)
                ws2 = wb['12월 AR(이수임)']

                count = 2
                arCount = 0
                ar_id_Count = 0
                totCount = 13
                
                # ars = AR.ARs()

                # results = []

                while count <= totCount:

                    cell = ws2.cell(row=count, column=4)
                    # CostCenter        = []
                    # Project           = []
                    # UnitPrice         = []
                    # InvoiceDetails    = []
                    # ids = AR.InvoiceDetails()

                    with open('./conf/template_InvoiceDetails.json') as template_InvoiceDetails:
                        json_ID = json.load(template_InvoiceDetails)

                    if isinstance(cell, MergedCell):
                        # print("{}) merged({}) : {} , {}".format(count, mergedCount, type(cell) == MergedCell, cell.value)) # Merged : True

                        mergedCount = mergedCount + 1

                        # Invoice Details
                        json_ID['MemoLine']          = ws2["I{}".format(count)].value
                        json_ID['Description']       = Description
                        json_ID['Quantity']          = ws2["K{}".format(count)].value
                        json_ID['UnitPrice']         = ws2["L{}".format(count)].value
                        json_ID['TaxClassification'] = 'HX_매출과세'
                        json_ID['TBC']               = 'Sales Transaction' # Transaction Business Category
                        # json_ID['Revenu']            = '' # 자동생성(MemoLine 추가하면 자동으로 등록됨)

                        # Distributions
                        json_ID['Company']           = 'HX'    # 자동생성
                        json_ID['Business']          = '999'   # 자동생성
                        json_ID['RESPCenter']        = '99999' # 자동생성
                        json_ID['CostCenter']        = ws2["P{}".format(count)].value
                        json_ID['Account']           = ws2["Q{}".format(count)].value
                        # SubAccount        = '' # 자동생성	
                        json_ID['Project']           = ws2["R{}".format(count)].value
                        # Product           = '' # 안건드림
                        # TBD               = '' # 안건드림

                        InvoiceDetails.append(json_ID)

                    else:
                        # print("{}) merged({}) : {}, {}".format(count, mergedCount, type(cell) == MergedCell, cell.value)) # Merged : False
                        ar  = AR.AR()
                        mergedCount = 1 # default(Merge를 안했더라도 기본값 1 설정)
                        ar_id_Count = 0

                        with open('./conf/template_GeneralInformation.json') as template_GeneralInformation:
                            json_GI = json.load(template_GeneralInformation)

                        # General Information
                        json_GI['BusinessUnit']      = 'HX_BU'
                        json_GI['TransactionSource'] = 'HX_MANUAL'
                        json_GI['TransactionType']   = '영업매출'
                        json_GI['TransactionNumber'] = ws2["D{}".format(count)].value # 생성로직 구현(기존에 만들어놨던거 가져오기)
                        # json_GI['TransactionNumber'] = 'ART{}{}{}'.format(str(datetime.now().strftime('%y%m%d')), '테스트', str(datetime.now().strftime('%H%M'))) # ARTyymmdd이름hhmm
                        json_GI['TransactionDate']   = datetime.now().strftime('%Y') + '-' + datetime.now().strftime('%m') + '-' + str(ws2["A{}".format(count)].value)
                        json_GI['AccountingDate']    = datetime.now().strftime('%Y') + '-' + datetime.now().strftime('%m') + '-' + str(ws2["A{}".format(count)].value)

                        # Customer
                        json_GI['BillToName']        = ws2["B{}".format(count)].value
                        json_GI['PaymentTerms']      = ws2["E{}".format(count)].value
                        json_GI['StructuredPaymentReference']= (ws2["F{}".format(count)].value).format(datetime.now().strftime('%m'))
                        json_GI['TaxProofType']      = '10(세금계산서[전자])' # 세무 증빙 유형	
                        json_GI['BusinessNumber']    = '0000' # 종사업장 번호	
                        json_GI['ReverseIssue']      = ws2["H{}".format(count)].value # 역발행여부(공급받는자가 계산서 내용을 작성해서 보내면 공급자가 확인 후 발행처리)
                        # Writer            = writer # 작성자 
                        json_GI['Email']             = account # 담당자 정보에서 가져오기(로그인 계정)
                        json_GI['RegisterYN']        = ws2["S{}".format(count)].value
                        InvoiceDetails    = []


                        # Invoice Details
                        json_ID['MemoLine']          = ws2["I{}".format(count)].value
                        json_ID['Description']       = (ws2["F{}".format(count)].value).format(datetime.now().strftime('%m'))
                        Description                  = (ws2["F{}".format(count)].value).format(datetime.now().strftime('%m'))
                        json_ID['Quantity']          = ws2["K{}".format(count)].value
                        json_ID['UnitPrice']         = ws2["L{}".format(count)].value
                        json_ID['TaxClassification'] = 'HX_매출과세'
                        json_ID['TBC']               = 'Sales Transaction' # Transaction Business Category
                        # json_ID['Revenu']            = '' # 자동생성(MemoLine 추가하면 자동으로 등록됨)

                        # Distributions
                        json_ID['Company']           = 'HX'    # 자동생성
                        json_ID['Business']          = '999'   # 자동생성
                        json_ID['RESPCenter']        = '99999' # 자동생성
                        json_ID['CostCenter']        = ws2["P{}".format(count)].value
                        json_ID['Account']           = ws2["Q{}".format(count)].value
                        # SubAccount        = '' # 자동생성	
                        json_ID['Project']           = ws2["R{}".format(count)].value
                        # Product           = '' # 안건드림
                        # TBD               = '' # 안건드림

                        InvoiceDetails.append(json_ID)

                        arCount = arCount + 1

                        logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                        logger.LoggerFactory._LOGGER.info('{} : json 생성 완료'.format(json_GI['TransactionNumber']))
                        # logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                    
                    json_GI['InvoiceDetails']    = InvoiceDetails
                    
                    # backup 폴더 생성
                    if (os.path.exists(backup) == False) :
                        os.makedirs(backup)

                    with open((backup + json_GI['TransactionNumber'] + '.json'), 'w') as f:
                        json.dump(json_GI, f, ensure_ascii=False)
                        
                    count = count + 1
                
                logger.LoggerFactory._LOGGER.info('--------------------------------------------------')

            except Exception as e:
                print('else - 2nd except')
                logger.LoggerFactory._LOGGER.info('error msg : {}'.format(e))
                logger.LoggerFactory._LOGGER.info('2) 엑셀에 등록된 데이터 기반 json 데이터 생성 <<< 실패')
                re_count = re_count + 1

            else:
                print('else - 2nd else')

                # try-except-else 내 다시 try문 사용가능 x 3
                try:
                    print('else - else - 3rd try')
                    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                    logger.LoggerFactory._LOGGER.info('1) 설치여부 확인 및 사용자 정보 가져오기')
                    logger.LoggerFactory._LOGGER.info('2) 엑셀에 등록된 데이터 기반 json 데이터 생성')
                    logger.LoggerFactory._LOGGER.info('3) ERP 로그인 및 Receivables > Billing 이동 <<< [현재위치]')
                    logger.LoggerFactory._LOGGER.info('4) Transaction 생성 및 General Information 등록')
                    logger.LoggerFactory._LOGGER.info('5) Invoice Lines 및 Distribution 등록')

                    # dev
                    # url = 'https://efuw-test.fa.ap1.oraclecloud.com/'

                    # live
                    url = 'https://efuw.login.ap1.oraclecloud.com/'

                    if displayYN == 'N':                    
                        # 1) 화면출력 안함
                        options = webdriver.ChromeOptions()
                        # options.add_experimental_option("detach", True) # 화면꺼짐 방지
                        options.add_argument('headless')
                        driver = webdriver.Chrome(options=options)
                        wait = WebDriverWait(driver, 20)
                        driver.get(url)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('ERP 페이지 이동 완료(화면출력 안함)')

                    else:
                        # 2) 화면출력 함
                        # 2.1)
                        # options = webdriver.ChromeOptions()
                        # options.add_experimental_option("detach", True) # 화면꺼짐 방지
                        # driver = webdriver.Chrome(options=options)
                        # 2.2) disconnected: not connected to DevTools >>> 해결안됨
                        # options = webdriver.ChromeOptions()
                        # options.add_argument('--no-sandbox')
                        # options.add_argument('--disable-dev-shm-usage')
                        # driver = webdriver.Chrome(options=options)
                        # The --no-sandbox keeps chromedriver from running in a sandbox process and the --disable-dev-shm-usage forces chromedriver to use the /tmp directory instead of /dev/shm because sometimes that can be too small for the website you are accessing.
                        # 2.3) 사용중
                        driver = webdriver.Chrome()
                        wait = WebDriverWait(driver, 20)

                        driver.get(url)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('ERP 페이지 이동 완료(화면출력 함)')

                    # 로그인
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnActive"]')))
                    displayOK = driver.find_element(By.XPATH, '//*[@id="btnActive"]').is_displayed()
                    logger.LoggerFactory._LOGGER.info('displayOK : {}'.format(displayOK))
                    # driver.find_element(By.XPATH, '//*[@id="userid"]').send_keys(account)
                    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div/main/form/input[1]').send_keys(account)
                    time.sleep(1)
                    driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(password)
                    time.sleep(1)
                    driver.find_element(By.XPATH, '//*[@id="btnActive"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('로그인 완료')
                    logger.LoggerFactory._LOGGER.info('{} : {}'.format(account, password))

                    logger.LoggerFactory._LOGGER.info('AR Transaction 이동')
                    # 햄버거 버튼 클릭
                    time.sleep(10)
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pt1:_UISmmLink::icon"]'))) # dev
                    # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="groupNode_my_information"]'))) # live - 화면이 모두 로딩되고 나서
                    driver.find_element(By.XPATH, '//*[@id="pt1:_UISmmLink::icon"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('햄버거 버튼 클릭 완료')

                    # Receivables 메뉴 클릭
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pt1:_UISnvr:0:nvgpgl2_groupNode_receivables"]')))
                    driver.find_element(By.XPATH, '//*[@id="pt1:_UISnvr:0:nvgpgl2_groupNode_receivables"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Receivables 메뉴 클릭 완료')

                    # Billing 클릭
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pt1:_UISnvr:0:nv_itemNode_receivables_billing"]')))
                    driver.find_element(By.XPATH, '//*[@id="pt1:_UISnvr:0:nv_itemNode_receivables_billing"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Billing 클릭 완료')


                except Exception as e:
                    print('else - else - 3rd except')
                    logger.LoggerFactory._LOGGER.info('error msg : {}'.format(e))
                    logger.LoggerFactory._LOGGER.info('3) ERP 로그인 및 Receivables > Billing 이동 <<< 실패')
                    logger.LoggerFactory._LOGGER.info('사용자 정보를 다시 확인해주시기 바랍니다.')
                    re_count = re_count + 1

                else:
                    print('else - else - 3rd else')

                    # try-except-else 내 다시 try문 사용가능 x 4
                    try:
                        print('else - else - else - 4th try')
                        logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                        logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                        logger.LoggerFactory._LOGGER.info('1) 설치여부 확인 및 사용자 정보 가져오기')
                        logger.LoggerFactory._LOGGER.info('2) 엑셀에 등록된 데이터 기반 json 데이터 생성')
                        logger.LoggerFactory._LOGGER.info('3) ERP 로그인 및 Receivables > Billing 이동')
                        logger.LoggerFactory._LOGGER.info('4) Transaction 생성 및 General Information 등록 <<< [현재위치]')
                        logger.LoggerFactory._LOGGER.info('5) Invoice Lines 및 Distribution 등록')

                        # fileList = os.listdir(backup)
                        fileList = natsort.natsorted(os.listdir(backup)) # 파일명 정렬
                        fileCount = len(fileList)
                        arCount = 0

                        if fileCount == 0:
                            raise FileNotFoundError

                        else:

                            ar_fst_row    = 2

                            while arCount < fileCount:

                                with open(backup + fileList[arCount]) as f:
                                    ar = json.load(f)
                                
                                # TransactionNumber    = ar['TransactionNumber']
                                # InvoiceDetails       = ar['InvoiceDetails']
                                # print(TransactionNumber)
                                # print(len(InvoiceDetails))
                                # print(InvoiceDetails[arCount]['MemoLine'])

                                # issue_cnt, not_issue_cnt, re_count
                                issue_cnt     = 0
                                not_issue_cnt = 0
                                # re_count      = 1 # default값 1회
                                # ar_fst_row    = 2

                                if ar['RegisterYN'] == 'Y':
                                    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                                    logger.LoggerFactory._LOGGER.info('{} : 기발행된 AR 입니다.'.format(ar['TransactionNumber']))
                                    # ar_fst_row = ar_fst_row + len(ar['InvoiceDetails'])
                                    print(ar_fst_row)
                                    print(len(ar['InvoiceDetails']))
                                    ar_fst_row = ar_fst_row + len(ar['InvoiceDetails'])
                                    print(ar_fst_row)

                                elif ar['RegisterYN'] == 'E':
                                    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                                    logger.LoggerFactory._LOGGER.info('{} : 기발행 오류가 발생한 AR 입니다.'.format(ar['TransactionNumber'])) 
                                    ar_fst_row = ar_fst_row + len(ar['InvoiceDetails'])

                                elif ar['RegisterYN'] == 'N':
                                    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                                    logger.LoggerFactory._LOGGER.info('{} : AR 발행 시작!!'.format(ar['TransactionNumber'])) 

                                    
                                    # Exception 관리하기 위해
                                    # GI_RegisterYN = 'N' : 저장이 안된 상태로 화면 리프레쉬
                                    # GI_RegisterYN = 'Y' : 저장이 된 상태로 등록된 AR 삭제
                                    GI_RegisterYN = 'N'

                                    # Tasks 이미지 클릭
                                    wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdi__TransactionsWorkArea_itemNode__FndTasksList::icon"]')))
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdi__TransactionsWorkArea_itemNode__FndTasksList::icon"]').click()
                                    # time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Tasks 이미지 클릭 완료')

                                    # Create Transaction 클릭
                                    wait.until(
                                        EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaT:0:RAtl1"]')))
                                    driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaT:0:RAtl1"]').click()
                                    logger.LoggerFactory._LOGGER.info('Create Transaction 클릭 완료')

                                    # 데이터 입력
                                    # Transaction Source
                                    wait.until(EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:batchSourceId::content"]')))
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:batchSourceId::content"]').send_keys(
                                        ar['TransactionSource'])
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:batchSourceId::content"]').send_keys(
                                        Keys.ENTER)
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Transaction Source 입력 완료')

                                    # Transaction Type : 매출유형
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:transactionTypeId::content"]').send_keys(
                                        ar['TransactionType'])
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:transactionTypeId::content"]').send_keys(
                                        Keys.ENTER)
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Transaction Type : 매출유형 입력 완료')

                                    # Transaction Number : 전표번호(ARTyymmdd이름hhmm)
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:inputText2::content"]').send_keys(
                                        ar['TransactionNumber'])
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Transaction Number : 전표번호 입력 완료')

                                    # Transaction Date : 계산서 발행일자(yyyy-mm-10) - 사용익월 10일
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:tdt::content"]').clear()
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:tdt::content"]').send_keys(
                                        ar['TransactionDate'])
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Transaction Date : AR 작성일자 입력 완료')

                                    # Accounting Date : 계산서 발행일자(yyyy-mm-10) - 사용익월 10일
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:inputDate9::content"]').clear()
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:inputDate9::content"]').send_keys(
                                        ar['TransactionDate'])
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Accounting Date : 계산서 발행일자 입력 완료')

                                    # Bill-to Name
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:billToNameId::content"]').send_keys(
                                        ar['BillToName'])
                                    # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:billToNameId::content"]').send_keys(Keys.ENTER)
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Bill-to Name	입력 완료')

                                    # Payment Terms	: 수금조건
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:paymentTermId::content"]').send_keys(
                                        ar['PaymentTerms'])
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:paymentTermId::content"]').send_keys(
                                        Keys.ENTER)
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Payment Terms : 수금조건 입력 완료')

                                    # Show More 클릭
                                    wait.until(EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showMore"]')))
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showMore"]').click()
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Show More 클릭 완료')

                                    # Miscellaneous 클릭
                                    wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showDetailItem5::disAcr"]')))
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showDetailItem5::disAcr"]').click()
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Miscellaneous 클릭 완료')

                                    # Structured Payment Reference : 적요 입력 필수
                                    wait.until(EC.element_to_be_clickable(
                                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:it20::content"]')))
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:it20::content"]').send_keys(
                                        ar['StructuredPaymentReference'])
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Structured Payment Reference : 적요 입력 완료')

                                    # 세무 증빙 유형(TaxProofType)	
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:df1_TransactionHeaderDFF2IteratorslipTypeHXSlipType::content"]').send_keys(
                                        ar['TaxProofType'])
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('세무 증빙 유형 입력 완료')

                                    # 역발행여부(ReverseIssue)
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:df1_TransactionHeaderDFF2IteratorinverseIssueHXSlipType::content"]').send_keys(
                                        ar['ReverseIssue'])
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('역발행여부 입력 완료')

                                    # 작성자 E-Mail
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:df1_TransactionHeaderDFF2IteratoruserEmailHXSlipType::content"]').send_keys(
                                        ar['Email'])
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('작성자 E-Mail 입력 완료')


                                    ids_count = 0

                                    # Invoice Lines
                                    while ids_count < len(ar['InvoiceDetails']):
                                        logger.LoggerFactory._LOGGER.info('{}번째 ids_count'.format(ids_count + 1))

                                        # Memo Line
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:memoLineNameId::content"]'.format(ids_count)).send_keys(
                                            ar['InvoiceDetails'][ids_count]['MemoLine'])
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:memoLineNameId::content"]'.format(ids_count)).send_keys(
                                            Keys.ENTER)
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Memo Line 입력 완료')

                                        # Description
                                        # Memo Line보다 먼저 입력하면 나중에 Memo Line 내용으로 바뀜
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:descriptionId::content"]'.format(ids_count)).clear()
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:descriptionId::content"]'.format(ids_count)).send_keys(
                                            ar['InvoiceDetails'][ids_count]['Description'])
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Description 입력 완료')

                                        # Quantity
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:quantity::content"]'.format(ids_count)).send_keys(
                                            ar['InvoiceDetails'][ids_count]['Quantity'])
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Quantity 완료')

                                        # Unit Price
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:sellingPrice::content"]'.format(ids_count)).send_keys(
                                            ar['InvoiceDetails'][ids_count]['UnitPrice'])
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Unit Price 완료')

                                        # Tax Classification : 과세유형
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:taxClassificationCodeId::content"]'.format(ids_count)).send_keys(
                                            ar['InvoiceDetails'][ids_count]['TaxClassification'])
                                        driver.find_element(By.XPATH,
                                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:taxClassificationCodeId::content"]'.format(ids_count)).send_keys(
                                            Keys.ENTER)
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Tax Classification 완료')

                                        ids_count = ids_count + 1


                                    # save
                                    wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:saveMenu"]/table/tbody/tr/td[1]/a/span')))
                                    # time.sleep(30)
                                    driver.find_element(By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:saveMenu"]/table/tbody/tr/td[1]/a/span').click()
                                    time.sleep(2)
                                    logger.LoggerFactory._LOGGER.info('Save 완료')

                                    # 만약 AR 문서가 등록되어 있는 경우 삭제프로세스 진행
                                    if driver.find_element(By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]').is_displayed(): 
                                        logger.LoggerFactory._LOGGER.info('AR 문서가 등록된 상태')

                                        # 오류 AR 삭제프로세스
                                        # 1) Billing 화면으로 이동 : refresh
                                        # 2) AR 검색화면으로 이동 : 돋보기 클릭
                                        # 3) Transaction Number 입력
                                        # 4) Transaction Number Search
                                        # 5) 해당 AR 클릭
                                        # 6) 삭제
                                        # 7) Done

                                        # This transaction number already exists. Enter a unique transaction number. (AR-855040)
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]')))
                                        driver.find_element(By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("1 - This transaction number already exists. Enter a unique transaction number. (AR-855040)")

                                        # Cancel
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a')))
                                        driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("1 - Cancel")

                                        # Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue?
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]')))
                                        driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue? >>> Yes 클릭 완료")

                                        # 1) Billing 화면으로 이동
                                        # driver.refresh()
                                        # time.sleep(2)
                                        # logger.LoggerFactory._LOGGER.info("refresh")

                                        # 2) AR 검색화면으로 이동 : 돋보기 클릭
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdiTransactionsQuickSearch::icon"]')))
                                        driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdiTransactionsQuickSearch::icon"]').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("2")

                                        # 3) Transaction Number 입력
                                        driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1:value00::content"]').clear()
                                        driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1:value00::content"]').send_keys(ar['TransactionNumber'])
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("3")

                                        # 4) Transaction Number Search
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1::search"]')))
                                        driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1::search"]').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("4")

                                        # 5) 해당 AR 클릭
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[2]/div/div/div/div/div/div/div/span/div[2]/div[2]/div/div[2]/table/tbody/tr/td[3]/span/a')))
                                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[2]/div/div/div/div/div/div/div/span/div[2]/div[2]/div/div[2]/table/tbody/tr/td[3]/span/a').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("5")

                                        # 6) 삭제
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[8]/div/a')))
                                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[8]/div/a').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("6")

                                        # You're about to delete incomplete transaction. Do you want to continue?
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]')))
                                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("You're about to delete incomplete transaction. Do you want to continue? >>> Yes 클릭 완료")

                                        # 7) Done
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button')))
                                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info("7")

                                        continue
                                        
                                    else: 
                                        logger.LoggerFactory._LOGGER.info('AR 문서가 등록안된 상태')

                                        # GI_RegisterYN = 'Y'

                                        # Action → Edit Distribution
                                        time.sleep(2)
                                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:3:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a
                                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a
                                        # /html/body/div[2]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[1]/div/div[1]/div[1]/table/tbody/tr/td/div/div/div/div/table/tbody/tr/td[2]/a
                                        # /html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[1]/div/div[1]/div[1]/table/tbody/tr/td/div/div/div/div/table/tbody/tr/td[2]/a
                                        wait.until(EC.element_to_be_clickable((By.XPATH,
                                                                            '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[1]/div/div[1]/div[1]/table/tbody/tr/td/div/div/div/div/table/tbody/tr/td[2]/a')))
                                        driver.find_element(By.XPATH,
                                                            '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[1]/div/div[1]/div[1]/table/tbody/tr/td/div/div/div/div/table/tbody/tr/td[2]/a').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Action 버튼 클릭 완료')
                                        # wait.until(EC.element_to_be_clickable((By.XPATH,
                                        #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:3:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a')))
                                        # driver.find_element(By.XPATH,
                                        #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:3:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a').click()
                                        # time.sleep(2)
                                        # logger.LoggerFactory._LOGGER.info('Action 버튼 클릭 완료')
                                        # wait.until(EC.element_to_be_clickable((By.XPATH,
                                        #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a')))
                                        # driver.find_element(By.XPATH,
                                        #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a').click()
                                        # time.sleep(2)
                                        # logger.LoggerFactory._LOGGER.info('Action 버튼 클릭 완료')

                                        # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:cmi7"]/td[2]').click()
                                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:cmi7"]/td[2]
                                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:cmi7"]/td[2]
                                        # /html/body/div[1]/form/div[2]/div[2]/div/div/div/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[1]/td[2]
                                        wait.until(EC.element_to_be_clickable(
                                            (By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div/div/div/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[1]/td[2]')))
                                        driver.find_element(By.XPATH,
                                                            '/html/body/div[1]/form/div[2]/div[2]/div/div/div/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[1]/td[2]').click()
                                        # wait.until(EC.element_to_be_clickable(
                                        #     (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi7"]/td[2]')))
                                        # driver.find_element(By.XPATH,
                                        #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi7"]/td[2]').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Edit Distribution 클릭 완료')

                                        # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:cb7"]').click()
                                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:cb7"]
                                        wait.until(EC.element_to_be_clickable(
                                            (By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]')))
                                        driver.find_element(By.XPATH,
                                                            '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]').click()
                                        # wait.until(EC.element_to_be_clickable(
                                        #     (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb7"]')))
                                        # driver.find_element(By.XPATH,
                                        #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb7"]').click()
                                        time.sleep(2)
                                        logger.LoggerFactory._LOGGER.info('Alert창 Yes 클릭 완료')
                                        
                                        GI_RegisterYN = 'Y'

                                        ids_count2 = 0
                                        # while ids_count2 < len(ars.result[issue_cnt].InvoiceDetails):
                                        while ids_count2 < len(InvoiceDetails):

                                            time.sleep(2)

                                            # Line Number
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[2]/span/input').clear()
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[2]/span/input').send_keys(ids_count2 + 1)
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Edit Distributions > Line Number')

                                            # Account Class
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input').clear()
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input').send_keys('Revenue')
                                            time.sleep(2)
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input').send_keys(Keys.ENTER)
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Edit Distributions > Account Class')

                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[2]/table/tbody/tr/td[6]/span/span/div/table/tbody/tr/td[3]/a/img')))
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[2]/table/tbody/tr/td[6]/span/span/div/table/tbody/tr/td[3]/a/img').click()
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Edit Distributions > 돋보기 버튼 클릭 완료')

                                            
                                            # Edit Distributions - Distributions - RESP Center
                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/span/span/input')))
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/span/span/input').clear()
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/span/span/input').send_keys('99999')
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Distributions > RESP Center 입력 완료')

                                            # Edit Distributions - Distributions - RESP Center 클릭
                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]')))
                                            time.sleep(1)
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]').click()
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Distributions > RESP Center 클릭 완료')

                                            # Edit Distributions - Distributions - Cost Center 입력
                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[5]/td[2]/table/tbody/tr/td[1]/span/span/input')))
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[5]/td[2]/table/tbody/tr/td[1]/span/span/input').clear()
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[5]/td[2]/table/tbody/tr/td[1]/span/span/input').send_keys(ar['InvoiceDetails'][ids_count2]['CostCenter'])
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Distributions > Cost Center 입력 완료')

                                            # Edit Distributions - Distributions - Cost Center 클릭
                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]')))
                                            time.sleep(1)
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]').click()
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Distributions > Cost Center 클릭 완료')

                                            # Edit Distributions - Distributions - Project 입력
                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[8]/td[2]/table/tbody/tr/td[1]/span/span/input')))
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[8]/td[2]/table/tbody/tr/td[1]/span/span/input').clear()
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[8]/td[2]/table/tbody/tr/td[1]/span/span/input').send_keys(ar['InvoiceDetails'][ids_count2]['Project'])  
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Distributions > Project 입력 완료')

                                            # Edit Distributions - Distributions - Project 클릭
                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]')))
                                            time.sleep(1)
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]').click()
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Distributions > Project 클릭 완료')


                                            # Edit Distributions - Distributions - OK 버튼 클릭
                                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/span/button[3]')))
                                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/span/button[3]').click()
                                            
                                            time.sleep(2)
                                            logger.LoggerFactory._LOGGER.info('Distributions > OK 버튼 클릭 완료')

                                            ids_count2 = ids_count2 + 1



                                        # Save and Close
                                        wait.until(EC.element_to_be_clickable(
                                            (By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]')))
                                        driver.find_element(By.XPATH,
                                                            '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]').click()
                                        time.sleep(3)  # 없애면 오류발생
                                        # wait.until(EC.element_to_be_clickable(
                                        #     (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb5"]')))
                                        # driver.find_element(By.XPATH,
                                        #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb5"]').click()
                                        # time.sleep(3)  # 없애면 오류발생

                                        # Save 우측 화살표버튼 클릭
                                        wait.until(EC.element_to_be_clickable(
                                            (By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[5]/div/table/tbody/tr/td[2]/div/a')))
                                        driver.find_element(By.XPATH,
                                                            '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[5]/div/table/tbody/tr/td[2]/div/a').click()
                                        time.sleep(3)  # 없애면 오류발생
                                        logger.LoggerFactory._LOGGER.info('Save 우측 화살표버튼 클릭 완료')
                                        # wait.until(EC.element_to_be_clickable(
                                        #     (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:saveMenu::popEl"]')))
                                        # driver.find_element(By.XPATH,
                                        #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:saveMenu::popEl"]').click()
                                        # time.sleep(3)  # 없애면 오류발생
                                        # logger.LoggerFactory._LOGGER.info('Save 우측 화살표버튼 클릭 완료')

                                        # Save and Close
                                        wait.until(EC.element_to_be_clickable(
                                            (By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div/div/div/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]')))
                                        driver.find_element(By.XPATH,
                                                            '/html/body/div[1]/form/div[2]/div[2]/div/div/div/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]').click()
                                        time.sleep(3)  # 없애면 오류발생
                                        logger.LoggerFactory._LOGGER.info('Save and Close 클릭 완료')
                                        # wait.until(EC.element_to_be_clickable(
                                        #     (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi10"]/td[2]')))
                                        # driver.find_element(By.XPATH,
                                        #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi10"]/td[2]').click()
                                        # time.sleep(3)  # 없애면 오류발생
                                        # logger.LoggerFactory._LOGGER.info('Save and Close 클릭 완료')

                                        # Billing 화면에서 Alert창 내 OK버튼 클릭
                                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div/div[1]/table/tbody/tr/td/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button')))
                                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div/div[1]/table/tbody/tr/td/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button').click()
                                        time.sleep(3)  # 없애면 오류발생
                                        logger.LoggerFactory._LOGGER.info('Billing 화면에서 Alert창 내 OK버튼 클릭 완료')
                                        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]')))
                                        # driver.find_element(By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]').click()
                                        # time.sleep(3)  # 없애면 오류발생
                                        # logger.LoggerFactory._LOGGER.info('Billing 화면에서 Alert창 내 OK버튼 클릭 완료')

                                        ws2['S{}'.format(ar_fst_row)] = 'Y'
                                        print(ar_fst_row)
                                        print(len(ar['InvoiceDetails']))
                                        ar_fst_row = ar_fst_row + len(ar['InvoiceDetails'])
                                        print(ar_fst_row)
                                        # ars.result[issue_cnt].RegisterYN = 'Y'
                                        logger.LoggerFactory._LOGGER.info('AR 발행완료 : {}'.format(ar['TransactionNumber']))
                                        wb.save(destination)

                                        # logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                                        # logger.LoggerFactory._LOGGER.info('{} : AR 발행완료.'.format(ar['TransactionNumber'])) 

                                        # arCount = arCount + 1
                                
                                else:
                                    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                                    logger.LoggerFactory._LOGGER.info('{} : RegisterYN 타입이 맞지 않은 AR 입니다.'.format(ar['TransactionNumber'])) 
                                    ar_fst_row = ar_fst_row + len(ar['InvoiceDetails'])
                                    # arCount = arCount + 1
                            
                                arCount = arCount + 1

                    except Exception as e:
                        print('else - else - else - 4th except')
                        logger.LoggerFactory._LOGGER.info('error msg : {}'.format(e))
                        logger.LoggerFactory._LOGGER.info('4) Transaction 생성 및 General Information 등록 <<< 실패')
                        logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                        logger.LoggerFactory._LOGGER.info('{} : 오류 발생으로 삭제 진행'.format(ar['TransactionNumber']))
                        logger.LoggerFactory._LOGGER.info('GI_RegisterYN : {}'.format(GI_RegisterYN))

                        # GI_RegisterYN = 'N' : General Information 저장 안된 상태(= Transaction Number가 등록안됨)
                        # GI_RegisterYN = 'Y' : General Information 저장 된 상태(= Transaction Number가 등록됨)
                        if GI_RegisterYN == 'N' : 
                            logger.LoggerFactory._LOGGER.info('General Information 저장 안된 상태(= Transaction Number가 등록안됨)')
                            time.sleep(2)

                            # 만약 AR 문서가 등록되어 있는 경우 삭제프로세스 진행
                            if driver.find_element(By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]').is_displayed(): 
                                logger.LoggerFactory._LOGGER.info('AR 문서가 등록되어 있는 경우 삭제프로세스 진행')

                                # This transaction number already exists. Enter a unique transaction number. (AR-855040)
                                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]')))
                                driver.find_element(By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("1 - This transaction number already exists. Enter a unique transaction number. (AR-855040)")

                                # Cancel
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton2"]/a')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton2"]/a').click()
                                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a')))
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("1 - Cancel")

                                # Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue?
                                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]')))
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue? >>> Yes 클릭 완료")

                                # 오류 AR 삭제프로세스
                                # 1) Billing 화면으로 이동 : refresh
                                # 2) AR 검색화면으로 이동 : 돋보기 클릭
                                # 3) Transaction Number 입력
                                # 4) Transaction Number Search
                                # 5) 해당 AR 클릭
                                # 6) 삭제
                                # 7) Done

                                # 1) Billing 화면으로 이동
                                # driver.refresh()
                                # time.sleep(2)
                                # logger.LoggerFactory._LOGGER.info("refresh")

                                # 2) AR 검색화면으로 이동 : 돋보기 클릭
                                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdiTransactionsQuickSearch::icon"]')))
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdiTransactionsQuickSearch::icon"]').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("2")

                                # 3) Transaction Number 입력
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1:value00::content"]').clear()
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1:value00::content"]').send_keys(ar['TransactionNumber'])
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("3")

                                # 4) Transaction Number Search
                                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1::search"]')))
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1::search"]').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("4")

                                # 5) 해당 AR 클릭
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]').click()
                                wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[2]/div/div/div/div/div/div/div/span/div[2]/div[2]/div/div[2]/table/tbody/tr/td[3]/span/a')))
                                driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[2]/div/div/div/div/div/div/div/span/div[2]/div[2]/div/div[2]/table/tbody/tr/td[3]/span/a').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("5")

                                # 6) 삭제
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]').click()
                                # time.sleep(2)

                                # Delete 버튼 클릭(저장 후)
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a').click()
                                wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[8]/div/a')))
                                driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[8]/div/a').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("6")

                                # You're about to delete incomplete transaction. Do you want to continue?
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]').click()
                                wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]')))
                                driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("You're about to delete incomplete transaction. Do you want to continue? >>> Yes 클릭 완료")

                                # 7) Done
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:cb1"]')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:cb1"]').click()
                                wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button')))
                                driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button').click()
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:MTF1:0:ap1:cb1"]')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:MTF1:0:ap1:cb1"]').click()
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/div/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button')))
                                # driver.find_element(By.XPATH, '/html/body/div/form/div/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("7")
                                
                                # 운영
                                # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:cb1"]
                                # /html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button

                                # 자동화 모듈
                                # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:MTF1:0:ap1:cb1"]
                                # /html/body/div/form/div/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button
                                

                            else:
                                logger.LoggerFactory._LOGGER.info('AR 문서가 등록되어 있는 경우 삭제프로세스 진행')

                                # Cancel 버튼 클릭(저장 전)
                                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a')))
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a').click()
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info('Cancel 버튼 클릭 완료')

                                # Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue?
                                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]')))
                                driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]').click()
                                # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]')))
                                # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]').click()
                                time.sleep(2)
                                logger.LoggerFactory._LOGGER.info("Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue? >>> Yes 클릭 완료")
                        
                            re_count = re_count + 1

                        else: 
                            logger.LoggerFactory._LOGGER.info('General Information 저장된 상태(= Transaction Number가 등록됨)')
                            
                            # 오류 AR 삭제프로세스
                            # 1) Billing 화면으로 이동 : refresh
                            # 2) AR 검색화면으로 이동 : 돋보기 클릭
                            # 3) Transaction Number 입력
                            # 4) Transaction Number Search
                            # 5) 해당 AR 클릭
                            # 6) 삭제
                            # 7) Done

                            # 1) Billing 화면으로 이동
                            driver.refresh()
                            time.sleep(2)
                            logger.LoggerFactory._LOGGER.info('1')

                            # # 2) AR 검색화면으로 이동 : 돋보기 클릭
                            # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdiTransactionsQuickSearch::icon"]')))
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdiTransactionsQuickSearch::icon"]').click()
                            # time.sleep(2)
                            # logger.LoggerFactory._LOGGER.info('2')

                            # # 3) Transaction Number 입력
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1:value00::content"]').clear()
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1:value00::content"]').send_keys(ar['TransactionNumber'])
                            # time.sleep(2)
                            # logger.LoggerFactory._LOGGER.info('3')

                            # # 4) Transaction Number Search
                            # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1::search"]')))
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaTj_id_1:1:qryId1::search"]').click()
                            # time.sleep(2)
                            # logger.LoggerFactory._LOGGER.info('4')

                            # # 5) 해당 AR 클릭
                            # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]')))
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]').click()
                            # time.sleep(2)
                            # logger.LoggerFactory._LOGGER.info('5')

                            # 6) 삭제
                            # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]')))
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]').click()
                            # time.sleep(2)

                            # Delete 버튼 클릭(저장 후)
                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[8]/div/a')))
                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr/td[8]/div/a').click()
                            # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a')))
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a').click()
                            time.sleep(2)
                            logger.LoggerFactory._LOGGER.info('6')

                            # You're about to delete incomplete transaction. Do you want to continue?
                            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]')))
                            driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/button[1]').click()
                            # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]')))
                            # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]').click()
                            time.sleep(2)
                            logger.LoggerFactory._LOGGER.info("You're about to delete incomplete transaction. Do you want to continue? >>> Yes 클릭 완료")

                            # # 7) Done
                            # # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]')))
                            # # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:AT2:_ATp:table2:0:cl1"]').click()
                            # # wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button')))
                            # # driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button').click()
                            # wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button')))
                            # driver.find_element(By.XPATH, '/html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button').click()
                            # # /html/body/div[1]/form/div[1]/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button
                            # # /html/body/div/form/div/div[1]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div[1]/div/div[2]/div/div/div/div/div/table/tbody/tr/td[2]/div/div/div[1]/div/div[1]/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/button
                            # # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:MTF1:0:ap1:cb1"]
                            # # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:MTF1:0:ap1:cb1"]
                            # time.sleep(10)
                            # logger.LoggerFactory._LOGGER.info("Done")

                        if re_count > 3 :
                            logger.LoggerFactory._LOGGER.info('AR 발행 중 오류 발생 3회 발생으로 에러처리 : {}'.format(ars.result[issue_cnt].BillToName))
                            ws2['S{}'.format(row_count)] = 'E'
                            ars.result[issue_cnt].RegisterYN = 'E'
                            wb.save(destination)

                        
                        driver.refresh()
                        time.sleep(10)

                        # # Cancel 버튼 여부(저장 전)
                        # if driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a').is_displayed() : 
                        #     # logger.LoggerFactory._LOGGER.info('햄버거 버튼 displayed : {}'.format(driver.find_element(By.XPATH, '//*[@id="_FOpt1:_UISmmLink::icon"]').is_displayed()))
                        #     driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:commandToolbarButton2"]/a').click()

                        #     # Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue?
                        #     wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]')))
                        #     driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:dialogCancel::yes"]').click()
                        #     time.sleep(2)
                        #     logger.LoggerFactory._LOGGER.info("Your changes aren''t saved. If you leave this page, then your changes will be lost. Do you want to continue? >>> Yes 클릭 완료")

                        # # Delete 버튼 여부(저장 후)
                        # elif driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a').is_displayed() : 
                        #     driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a').click()
                        #     # logger.LoggerFactory._LOGGER.info('햄버거 버튼 displayed : {}'.format(driver.find_element(By.XPATH, '//*[@id="_FOpt1:_UISmmLink::icon"]').is_displayed()))

                        #     # You're about to delete incomplete transaction. Do you want to continue?
                        #     wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]')))
                        #     driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]').click()
                        #     time.sleep(2)
                        #     logger.LoggerFactory._LOGGER.info("You're about to delete incomplete transaction. Do you want to continue? >>> Yes 클릭 완료")

                        # else :
                        #     logger.LoggerFactory._LOGGER.info('햄버거 버튼 displayed : {}'.format(driver.find_element(By.XPATH, '//*[@id="_FOpt1:_UISmmLink::icon"]').is_displayed()))
                        #     # driver.refresh()
                        #     logger.LoggerFactory._LOGGER.info('화면 refresh')


                        re_count = re_count + 1

                        driver.quit() # or close()

                        processing_ERP_AR_Transaction(0, 0, 0, re_count)

                    else:
                        print('else - else - else - 4th else')

                        # try-except-else 내 다시 try문 사용가능 x 5
                        try:
                            print('else - else - else - 5th try')

                        except Exception as e:
                            print('else - else - else - 5th except')
                            logger.LoggerFactory._LOGGER.info('error msg : {}'.format(e))
                            logger.LoggerFactory._LOGGER.info('5) Invoice Lines 및 Distribution 등록 <<< 실패')

                            driver.refresh()
                            logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
                            logger.LoggerFactory._LOGGER.info('현재 재시작 횟 수 : {}회'.format(re_count))
                            logger.LoggerFactory._LOGGER.info('{} : 오류 발생으로 삭제 진행'.format(ar['TransactionNumber']))

                            if re_count > 3 :
                                logger.LoggerFactory._LOGGER.info('AR 발행 중 오류 발생 3회 발생으로 에러처리 : {}'.format(ars.result[issue_cnt].BillToName))
                                ws2['S{}'.format(row_count)] = 'E'
                                ars.result[issue_cnt].RegisterYN = 'E'
                                wb.save(destination)
                            
                            # Delete 버튼 클릭
                            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a')))
                            driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a').click()
                            time.sleep(2)
                            logger.LoggerFactory._LOGGER.info('Delete 버튼 클릭 완료')

                            # You're about to delete incomplete transaction. Do you want to continue?
                            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]')))
                            driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]').click()
                            time.sleep(2)
                            logger.LoggerFactory._LOGGER.info("You're about to delete incomplete transaction. Do you want to continue? >>> Yes 클릭 완료")

                            re_count = re_count + 1

                            driver.quit() # or close()

                            processing_ERP_AR_Transaction(0, 0, 0, re_count)

                        else:
                            print('else - else - else - 5th else')

                        finally:
                            print('else - else - else - 5th finally')

                    finally:
                        print('else - else - else - 4th finally')

                finally:
                    print('else - else - 3rd finally')

            finally:
                print('else - 2nd finally')
        
        finally:
            print('1st finally')
            logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
            logger.LoggerFactory._LOGGER.info('현재 모듈 재구동 횟 수 : {}회'.format(re_count))
            logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
    
    else:
        logger.LoggerFactory._LOGGER.info('설치를 통해 사용자정보 등록 먼저 해주셔야합니다.')


processing_ERP_AR_Transaction(0, 0, 0, 1)