from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.edge import service
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support import expected_conditions as EC
import time

from webdriver_manager.core import driver

import logger
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from dto import AR
import json
# from flask import jsonify
import os

def processing_ERP_AR_Transaction(count, issue_cnt, not_issue_cnt, re_count):

    if count == 0:
        logger.LoggerFactory.create_logger()
    
    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
    logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
    logger.LoggerFactory._LOGGER.info('ERP AR 발행 시작')
    logger.LoggerFactory._LOGGER.info('1) 설치여부 확인 및 사용자 정보 가져오기')
    logger.LoggerFactory._LOGGER.info('2) ')
    logger.LoggerFactory._LOGGER.info('3) ')

    today = datetime.today()
    this_month = str(today.strftime('%m'))

    if this_month == '01':
        this_year = datetime(today.year, today.month, 1) + relativedelta(years=-1)
        this_month = datetime(today.year, today.month, 1) + relativedelta(months=-1)
    else :
        this_year = datetime(today.year, today.month, 1) + relativedelta(years=0)
        this_month = datetime(today.year, today.month, 1) + relativedelta(months=-1)

    year = str(today.strftime('%Y'))
    month = str(today.strftime('%m'))


    
    logger.LoggerFactory._LOGGER.info('설치된 사용자 정보 확인')
    with open('./conf/config.json') as f:
        user_config = json.load(f)

    installed   = user_config['installed']

    # 1) 설치여부 확인 및 사용자 정보 가져오기
    if installed == 'Y':

        user        = user_config['user']
        account     = user_config['account']
        password    = user_config['password']
        destination = user_config['filepath'] + user_config['filename']

        logger.LoggerFactory._LOGGER.info('user        = {}'.format(user))
        logger.LoggerFactory._LOGGER.info('account     = {}'.format(account))
        logger.LoggerFactory._LOGGER.info('password    = {}'.format(password))
        logger.LoggerFactory._LOGGER.info('installed   = {}'.format(installed))
        logger.LoggerFactory._LOGGER.info('destination = {}'.format(destination))

        # wb = load_workbook(destination)

        # # # ERP 페이지 이동(Chrome)
        # driver = webdriver.Chrome()
        # wait = WebDriverWait(driver, 20)

        # # # dev
        # # # url = 'https://efuw-test.fa.ap1.oraclecloud.com/'

        # # # live
        # url = 'https://efuw.login.ap1.oraclecloud.com/'

        # driver.get(url)
        # time.sleep(2)
        # logger.LoggerFactory._LOGGER.info('ERP 페이지 이동 완료')

        # # 로그인
        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnActive"]')))
        # displayOK = driver.find_element(By.XPATH, '//*[@id="btnActive"]').is_displayed()
        # logger.LoggerFactory._LOGGER.info('displayOK : {}'.format(displayOK))
        # # driver.find_element(By.XPATH, '//*[@id="userid"]').send_keys(account)
        # driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div/main/form/input[1]').send_keys(account)
        # time.sleep(1)
        # driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(password)
        # time.sleep(1)
        # driver.find_element(By.XPATH, '//*[@id="btnActive"]').click()
        # time.sleep(2)
        # logger.LoggerFactory._LOGGER.info('로그인 완료')
        # logger.LoggerFactory._LOGGER.info('{} : {}'.format(account, password))

    # else:
    #     logger.LoggerFactory._LOGGER.info('설치를 통해 사용자 정보 등록 부탁드립니다.')


        # AR Transaction 이동
        try:
            # wb = load_workbook(destination)

            # # ERP 페이지 이동(Chrome)
            driver = webdriver.Chrome()
            wait = WebDriverWait(driver, 20)

            # # dev
            # # url = 'https://efuw-test.fa.ap1.oraclecloud.com/'

            # # live
            url = 'https://efuw.login.ap1.oraclecloud.com/'

            driver.get(url)
            time.sleep(2)
            logger.LoggerFactory._LOGGER.info('ERP 페이지 이동 완료')

            # 로그인
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnActive"]')))
            displayOK = driver.find_element(By.XPATH, '//*[@id="btnActive"]').is_displayed()
            logger.LoggerFactory._LOGGER.info('displayOK : {}'.format(displayOK))
            # driver.find_element(By.XPATH, '//*[@id="userid"]').send_keys(account)
            driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div/main/form/input[1]').send_keys(account)
            time.sleep(1)
            driver.find_element(By.XPATH, '//*[@id="password"]').send_keys(password)
            time.sleep(1)
            driver.find_element(By.XPATH, '//*[@id="btnActive"]').click()
            time.sleep(2)
            logger.LoggerFactory._LOGGER.info('로그인 완료')
            logger.LoggerFactory._LOGGER.info('{} : {}'.format(account, password))

            logger.LoggerFactory._LOGGER.info('AR Transaction 이동')
            # 햄버거 버튼 클릭
            time.sleep(10)
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pt1:_UISmmLink::icon"]'))) # dev
            # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="groupNode_my_information"]'))) # live - 화면이 모두 로딩되고 나서
            driver.find_element(By.XPATH, '//*[@id="pt1:_UISmmLink::icon"]').click()
            time.sleep(2)
            logger.LoggerFactory._LOGGER.info('햄버거 버튼 클릭 완료')

            # Receivables 메뉴 클릭
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pt1:_UISnvr:0:nvgpgl2_groupNode_receivables"]')))
            driver.find_element(By.XPATH, '//*[@id="pt1:_UISnvr:0:nvgpgl2_groupNode_receivables"]').click()
            time.sleep(2)
            logger.LoggerFactory._LOGGER.info('Receivables 메뉴 클릭 완료')

            # Billing 클릭
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pt1:_UISnvr:0:nv_itemNode_receivables_billing"]')))
            driver.find_element(By.XPATH, '//*[@id="pt1:_UISnvr:0:nv_itemNode_receivables_billing"]').click()
            time.sleep(2)
            logger.LoggerFactory._LOGGER.info('Billing 클릭 완료')

        except Exception as ex:
            print('error msg : {}'.format(ex))
            # logger.LoggerFactory._LOGGER.info('전표 발행 대상 리스트 작성 실패 : {}'.format(BillToName))
            # logger.LoggerFactory._LOGGER.info('error msg : {}'.format(ex))
            # break
        




        # ★AR종합 가져오기
        # ws2 = wb['★AR종합']
        wb = load_workbook(destination)
        ws2 = wb['12월 AR(이수임)']
        # project_row = 5
        # print(ws2.merged_cells.ranges)
        # print(len(ws2.merged_cells.ranges))

        # General Information
        BusinessUnit      = 'HX_BU'
        TransactionSource = 'HX_MANUAL'
        TransactionType   = '영업매출'
        TransactionNumber = '' # 생성로직 구현(기존에 만들어놨던거 가져오기)
        TransactionDate   = '' # 값 가져오기
        AccountingDate    = '' # 값 가져오기

        # Customer
        BillToName        = '' # 값 가져오기
        PaymentTerms      = '' # 값 가져오기
        SPR               = '' # 값 가져오기, StructuredPaymentReference
        TaxProofType      = '10(세금계산서[전자])' # 세무 증빙 유형	
        BusinessNumber    = '0000' # 종사업장 번호	
        ReverseIssue      = '' # 역발행여부(공급받는자가 계산서 내용을 작성해서 보내면 공급자가 확인 후 발행처리)
        # Writer            = writer # 작성자 
        Email             = account # 담당자 정보에서 가져오기(로그인 계정)

        # Invoice Details
        InvoiceDetails    = []
        MemoLine          = '' # 값 가져오기
        Description       = '' # 값 가져오기
        Quantity          = '' # 값 가져오기
        UnitPrice         = [] # 값 가져오기
        TaxClassification = 'HX_매출과세'
        TBC               = '' # Transaction Business Category
        Revenu            = '' # 자동생성(MemoLine 추가하면 자동으로 등록됨)

        # Distributions
        Company           = 'HX'    # 자동생성
        Business          = '999'   # 자동생성
        RESPCenter        = '99999' # 자동생성
        CostCenter        = [] # 값 가져오기
        Account           = '' # 값 가져오기
        # SubAccount        = '' # 자동생성	
        Project           = [] # 값 가져오기
        # Product           = '' # 안건드림
        # TBD               = '' # 안건드림

        # count = 4
        # arCount = 0
        # totCount = 10
        count = 2
        arCount = 0
        totCount = 13

        ars = AR.ARs()

        while count <= totCount:
            cell = ws2.cell(row=count, column=4)
            CostCenter        = []
            Project           = []
            InvoiceDetails    = []
            UnitPrice         = []
            ids = AR.InvoiceDetails()

            # cell = ws2.cell(row=5, column=4)
            if isinstance(cell, MergedCell):
                # print("{}) merged({}) : {} , {}".format(count, mergedCount, type(cell) == MergedCell, cell.value)) # Merged : True
                # print(len(ars.result))
                # print(ars.result[0].CostCenter)
                # ars.result[arCount-1].CostCenter.append(ws2["E{}".format(count)].value)

                mergedCount = mergedCount + 1

                # Invoice Details
                ids.setMemoLine(ws2["I{}".format(count)].value)
                ids.setDescription(Description.format(datetime.now().strftime('%m'))) # Structured Payment Reference값 동일하게 사용
                ids.setQuantity(ws2["K{}".format(count)].value)
                if ws2["L{}".format(count)].value == None:  
                    ids.setUnitPrice(0)
                else: 
                    ids.setUnitPrice(ws2["L{}".format(count)].value)
                ids.setTaxClassification('HX_매출과세')
                ids.setTBC('Sales Transaction') # Transaction Business Category(*바뀔 수 있음)

                # Distributions
                ids.setCostCenter(ws2["P{}".format(count)].value)
                ids.setAccount(ws2["Q{}".format(count)].value)
                ids.setProject(ws2["R{}".format(count)].value)
                # InvoiceDetails.append(ids)
                ars.result[arCount-1].InvoiceDetails.append(ids)
            else:
                # print("{}) merged({}) : {}, {}".format(count, mergedCount, type(cell) == MergedCell, cell.value)) # Merged : False
                ar  = AR.AR()
                mergedCount = 1 # default(Merge를 안했더라도 기본값 1 설정)

                ar.setBusinessUnit('HX_BU')
                ar.setTransactionSource('HX_MANUAL') 
                ar.setTransactionType('영업매출') # 값 가져오기(*바뀔 수 있음)

                TransactionDate = datetime.now().strftime('%Y') + '-' + datetime.now().strftime('%m') + '-' + str(ws2["A{}".format(count)].value)
                ar.setTransactionDate(TransactionDate) 
                ar.setAccountingDate(TransactionDate) # TransactionDate와 동일
                # ar.setTransactionDate(ws2["A{}".format(count)].value) 
                # ar.setAccountingDate(ws2["D{}".format(count)].value) # TransactionDate와 동일

                ar.setTransactionNumber(ws2["D{}".format(count)].value)

                # logger.LoggerFactory._LOGGER.info('********** : {}'.format(str(ws2["D{}".format(count)].value)))
                # if str(ws2["D{}".format(count)].value) != None and str(ws2["D{}".format(count)].value) != '':    
                #     logger.LoggerFactory._LOGGER.info('********** if')
                #     ar.setTransactionNumber(ws2["D{}".format(count)].value)
                # else:
                #     logger.LoggerFactory._LOGGER.info('********** else')
                #     # 방법1)
                #     if arCount < 10 :
                #         AR_count = '0' + str(arCount)
                #     else :
                #         AR_count = str(arCount)
                #     # ar.setTransactionNumber('ART{}{}{}{}{}0{}'.format(str(datetime.now().strftime('%y')), str(datetime.now().strftime('%m')), str(ws2["A{}".format(count)].value), user, str(datetime.now().strftime('%H')), str(arCount))) # 생성로직 구현(기존에 만들어놨던거 가져오기)
                #     ar.setTransactionNumber('ART{}{}{}{}{}{}'.format(str(datetime.now().strftime('%y')), str(datetime.now().strftime('%m')), str(ws2["A{}".format(count)].value), user, str(datetime.now().strftime('%H')), AR_count))
                #     # 방법2)
                #     # ar.setTransactionNumber('ART{}{}{}{}{}{}'.format(str(datetime.now().strftime('%y')), str(datetime.now().strftime('%m')), str(ws2["A{}".format(count)].value), user, str(datetime.now().strftime('%H')), str(datetime.now().strftime('%M')))) 

                # Customer
                ar.setBillToName(ws2["B{}".format(count)].value) # ERP에 입력하는 업체명
                ar.setPaymentTerms(ws2["E{}".format(count)].value)
                ar.setStructuredPaymentReference((ws2["F{}".format(count)].value).format(datetime.now().strftime('%m'))) 
                ar.setTaxProofType('10(세금계산서[전자])') # 세무 증빙 유형	
                ar.setBusinessNumber('0000') # 종사업장 번호
                ar.setReverseIssue(ws2["H{}".format(count)].value) # 역발행여부(공급받는자가 계산서 내용을 작성해서 보내면 공급자가 확인 후 발행처리)
                # ar.setWriter('') # 작성자 
                ar.setEmail(account) # 담당자 정보에서 가져오기(로그인 계정)
                ar.setRegisterYN(ws2["S{}".format(count)].value) # AR 등록여부 확인

                # Invoice Details
                ar.setInvoiceDetails([]) 
                ids.setMemoLine(ws2["I{}".format(count)].value) 
                Description = ws2["F{}".format(count)].value
                ids.setDescription(Description.format(datetime.now().strftime('%m'))) # Structured Payment Reference값 동일하게 사용
                ids.setQuantity(ws2["K{}".format(count)].value)
                # ids.setQuantity('')
                if ws2["L{}".format(count)].value == None:  
                    ids.setUnitPrice(0)
                else: 
                    ids.setUnitPrice(ws2["L{}".format(count)].value)
                ids.setTaxClassification('HX_매출과세')
                ids.setTBC('Sales Transaction') # Transaction Business Category(*바뀔 수 있음)
                # ids.setRevenu('') # 자동생성(MemoLine 추가하면 자동으로 등록됨)

                # Distributions
                # ids.setCompany('HX') # 자동생성
                # ids.setBusiness('999') # 자동생성
                # ids.setRESPCenter('99999') # 자동생성
                ids.setCostCenter(ws2["P{}".format(count)].value)
                ids.setAccount(ws2["Q{}".format(count)].value)
                # ids.setSubAccount('') # 자동생성
                ids.setProject(ws2["R{}".format(count)].value)
                # ids.setProduct('') # 안건드림
                # ids.setTBD('') # 안건드림

                # CostCenter.append(ws2["E{}".format(count)].value)
                # ar.setCostCenter(CostCenter)
                
                # Project.append(ws2["G{}".format(count)].value)
                # ar.setProject(Project)
                # ids.setUnitPrice(ws2["J{}".format(count)].value)
                InvoiceDetails.append(ids)
                ar.setInvoiceDetails(InvoiceDetails)

                ars.setResult(ar)
                # logger.LoggerFactory._LOGGER.info('{}번째 AR : {}'.format(arCount, ar))

                arCount = arCount + 1
            
            # ars.setResult(ar)
            # backup 폴더 생성
            # if (os.path.exists('./backup/' + datetime.now().strftime('%Y%m%d')) == False) :
            #     os.makedirs('./backup/' + datetime.now().strftime('%Y%m%d'))
            if (os.path.exists('./backup/' + datetime.now().strftime('%Y') + '/' + datetime.now().strftime('%m') + '/' + datetime.now().strftime('%d') + '/' + datetime.now().strftime('%H%M') + '/') == False) :
                os.makedirs('./backup/' + datetime.now().strftime('%Y') + '/' + datetime.now().strftime('%m') + '/' + datetime.now().strftime('%d') + '/' + datetime.now().strftime('%H%M') + '/')

            # with open('./backup/{}/{}.json'.format(datetime.now().strftime('%Y%m%d'), ar.TransactionNumber), 'w') as fp:
            with open('./backup/{}/{}/{}/{}/{}.json'.format(datetime.now().strftime('%Y'), datetime.now().strftime('%m'), datetime.now().strftime('%d'), datetime.now().strftime('%H%M'), ar.TransactionNumber), 'w') as fp:
                json.dump(ars.toJsonARs(), fp, ensure_ascii=False)
            count = count + 1


        # AR발행 시작
        row = 2
        count = 0
        # print(len(ars.result))
        # logger.LoggerFactory._LOGGER.info('발행할 AR 갯수 : {}개'.format(len(ars.result)))

        while count < len(ars.result):
            # logger.LoggerFactory._LOGGER.info('발행할 AR 갯수 : {}개, {}번째 row'.format(len(ars.result), row))
            logger.LoggerFactory._LOGGER.info('##################################################')
            logger.LoggerFactory._LOGGER.info('##################################################')
            logger.LoggerFactory._LOGGER.info('발행할 AR 갯수 : {}개, {}번째 row, {}번째 배열'.format(len(ars.result), row, count))

            logger.LoggerFactory._LOGGER.info('General Information')
            logger.LoggerFactory._LOGGER.info('BusinessUnit      : {}'.format(ars.result[count].BusinessUnit))
            logger.LoggerFactory._LOGGER.info('TransactionSource : {}'.format(ars.result[count].TransactionSource))
            logger.LoggerFactory._LOGGER.info('TransactionType   : {}'.format(ars.result[count].TransactionType))
            logger.LoggerFactory._LOGGER.info('TransactionDate   : {}'.format(ars.result[count].TransactionDate))
            logger.LoggerFactory._LOGGER.info('AccountingDate    : {}'.format(ars.result[count].AccountingDate))
            logger.LoggerFactory._LOGGER.info('TransactionNumber : {}'.format(ars.result[count].TransactionNumber))
            logger.LoggerFactory._LOGGER.info('BillToName        : {}'.format(ars.result[count].BillToName))
            logger.LoggerFactory._LOGGER.info('PaymentTerms      : {}'.format(ars.result[count].PaymentTerms))
            logger.LoggerFactory._LOGGER.info('StructuredPaymentReference: {}'.format(ars.result[count].StructuredPaymentReference))
            logger.LoggerFactory._LOGGER.info('TaxProofType(세무증빙유형) : {}'.format(ars.result[count].TaxProofType))
            logger.LoggerFactory._LOGGER.info('BusinessNumber    : {}'.format(ars.result[count].BusinessNumber))
            logger.LoggerFactory._LOGGER.info('ReverseIssue      : {}'.format(ars.result[count].ReverseIssue))
            logger.LoggerFactory._LOGGER.info('Email             : {}'.format(ars.result[count].Email))

            # InvoiceDetails
            ids = 0
            for ids in range(len(ars.result[count].InvoiceDetails)):
                logger.LoggerFactory._LOGGER.info('##################################################')
                # logger.LoggerFactory._LOGGER.info(
                #     'Invoice Details : {}번째 row'.format(len(ars.result[count].InvoiceDetails)))
                logger.LoggerFactory._LOGGER.info('Invoice Details : {}번째'.format(ids + 1))
                logger.LoggerFactory._LOGGER.info('MemoLine        : {}'.format(ars.result[count].InvoiceDetails[ids].MemoLine))
                logger.LoggerFactory._LOGGER.info('Description     : {}'.format(ars.result[count].InvoiceDetails[ids].Description)) # Structured Payment Reference값 동일하게 사용
                logger.LoggerFactory._LOGGER.info('Quantity        : {}'.format(ars.result[count].InvoiceDetails[ids].Quantity))
                logger.LoggerFactory._LOGGER.info('UnitPrice       : {}'.format(ars.result[count].InvoiceDetails[ids].UnitPrice))
                logger.LoggerFactory._LOGGER.info('TaxClassification : {}'.format(ars.result[count].InvoiceDetails[ids].TaxClassification))
                logger.LoggerFactory._LOGGER.info('TBC             : {}'.format(ars.result[count].InvoiceDetails[ids].TBC))
                logger.LoggerFactory._LOGGER.info('CostCenter      : {}'.format(ars.result[count].InvoiceDetails[ids].CostCenter))
                logger.LoggerFactory._LOGGER.info('Account         : {}'.format(ars.result[count].InvoiceDetails[ids].Account))
                logger.LoggerFactory._LOGGER.info('Project         : {}'.format(ars.result[count].InvoiceDetails[ids].Project))

                ids = ids + 1


            row = row + 1
            count = count + 1


        logger.LoggerFactory._LOGGER.info('--------------------------------------------------')
        logger.LoggerFactory._LOGGER.info('ERP AR 발행 시작')
        logger.LoggerFactory._LOGGER.info('발행할 AR 갯수 : {}개'.format(len(ars.result)))

        # issue_cnt, not_issue_cnt, re_count
        issue_cnt     = 0
        not_issue_cnt = 0
        re_count      = 1 # default값 1회
        row_count     = 2

        while issue_cnt < len(ars.result):

            if ws2['S{}'.format(row_count)].value == 'E':
                logger.LoggerFactory._LOGGER.info('{}번쨰 : {} 기발행 실패(에러 or 사용자 없음)'.format(issue_cnt + 1, ars.result[issue_cnt].BillToName))
            elif ws2['S{}'.format(row_count)].value == 'X':
                logger.LoggerFactory._LOGGER.info('{}번쨰 : {} 발행 안함(선입금 or 선불)'.format(issue_cnt + 1, ars.result[issue_cnt].BillToName))
            elif ws2['S{}'.format(row_count)].value == 'N':
                logger.LoggerFactory._LOGGER.info('#################### {}번째 : {} ####################'.format(issue_cnt + 1, ars.result[issue_cnt].BillToName))
            
            logger.LoggerFactory._LOGGER.info('Invoice Details 갯수 : {}개'.format(len(ars.result[issue_cnt].InvoiceDetails)))
            # logger.LoggerFactory._LOGGER.info('{}번째 row'.format(row_count))

            ids_count = 0
            
            BillToName = ars.result[issue_cnt].BillToName
            logger.LoggerFactory._LOGGER.info('BillToName : [{}]'.format(BillToName))

            # if BillToName != '' and BillToName != None:
            if ars.result[issue_cnt].RegisterYN == 'N' : 

                # unit_price = ws2['D{}'.format(billStatement_row)].value
                # transaction_number = 'ART{}{}{}'.format(str(datetime.now().strftime('%y%m%d')), writer, str(datetime.now().strftime('%H%M'))) # ARTyymmdd이름hhmm
                # transaction_number = 'ART{}{}{}{}{}'.format(str(datetime.now().strftime('%y')),
                #                                             str(datetime.now().strftime('%m')),
                #                                             date_in_transaction_number, writer,
                #                                             str(datetime.now().strftime('%H%M')))  # ARTyymmdd이름hhmm


                # ws4 = wb['업체(청구단위)']
                # customer_row = 2
                # while customer_row <= 51:
                #     if BillToName == ws4['C{}'.format(customer_row)].value:
                #         bill_to_name = ws4['D{}'.format(customer_row)].value
                #         cust_sn = ws4['B{}'.format(customer_row)].value

                #     customer_row = customer_row + 1

                try:
                    # Tasks 이미지 클릭
                    wait.until(EC.element_to_be_clickable((By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdi__TransactionsWorkArea_itemNode__FndTasksList::icon"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTsdi__TransactionsWorkArea_itemNode__FndTasksList::icon"]').click()
                    # time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Tasks 이미지 클릭 완료')

                    # Create Transaction 클릭
                    wait.until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaT:0:RAtl1"]')))
                    driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:_FOTRaT:0:RAtl1"]').click()
                    logger.LoggerFactory._LOGGER.info('Create Transaction 클릭 완료')

                    # 데이터 입력
                    # Transaction Source
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:batchSourceId::content"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:batchSourceId::content"]').send_keys(
                        ars.result[issue_cnt].TransactionSource)
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:batchSourceId::content"]').send_keys(
                        Keys.ENTER)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Transaction Source 입력 완료')

                    # Transaction Type : 매출유형
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:transactionTypeId::content"]').send_keys(
                        ars.result[issue_cnt].TransactionType)
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:transactionTypeId::content"]').send_keys(
                        Keys.ENTER)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Transaction Type : 매출유형 입력 완료')

                    # Transaction Number : 전표번호(ARTyymmdd이름hhmm)
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:inputText2::content"]').send_keys(
                        ars.result[issue_cnt].TransactionNumber)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Transaction Number : 전표번호 입력 완료')

                    # Transaction Date : 계산서 발행일자(yyyy-mm-10) - 사용익월 10일
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:tdt::content"]').clear()
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:tdt::content"]').send_keys(
                        ars.result[issue_cnt].TransactionDate)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Transaction Date : AR 작성일자 입력 완료')

                    # Accounting Date : 계산서 발행일자(yyyy-mm-10) - 사용익월 10일
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:inputDate9::content"]').clear()
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:inputDate9::content"]').send_keys(
                        ars.result[issue_cnt].TransactionDate)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Accounting Date : 계산서 발행일자 입력 완료')

                    # Bill-to Name
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:billToNameId::content"]').send_keys(
                        ars.result[issue_cnt].BillToName)
                    # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:billToNameId::content"]').send_keys(Keys.ENTER)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Bill-to Name	입력 완료')

                    # Payment Terms	: 수금조건
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:paymentTermId::content"]').send_keys(
                        ars.result[issue_cnt].PaymentTerms)
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:paymentTermId::content"]').send_keys(
                        Keys.ENTER)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Payment Terms : 수금조건 입력 완료')

                    # Show More 클릭
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showMore"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showMore"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Show More 클릭 완료')

                    # Miscellaneous 클릭
                    wait.until(EC.element_to_be_clickable((By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showDetailItem5::disAcr"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:showDetailItem5::disAcr"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Miscellaneous 클릭 완료')

                    # Structured Payment Reference : 적요 입력 필수
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:it20::content"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:it20::content"]').send_keys(
                        ars.result[issue_cnt].StructuredPaymentReference)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Structured Payment Reference : 적요 입력 완료')

                    # 세무 증빙 유형(TaxProofType)	
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:df1_TransactionHeaderDFF2IteratorslipTypeHXSlipType::content"]').send_keys(
                        ars.result[issue_cnt].TaxProofType)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('세무 증빙 유형 입력 완료')

                    # 역발행여부(ReverseIssue)
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:df1_TransactionHeaderDFF2IteratorinverseIssueHXSlipType::content"]').send_keys(
                        ars.result[issue_cnt].ReverseIssue)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('세무 증빙 유형 입력 완료')

                    # 작성자 E-Mail
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:df1_TransactionHeaderDFF2IteratoruserEmailHXSlipType::content"]').send_keys(
                        ars.result[issue_cnt].Email)
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('작성자 E-Mail 입력 완료')



                    # Invoice Lines
                    while ids_count < len(ars.result[issue_cnt].InvoiceDetails):
                        logger.LoggerFactory._LOGGER.info('{}번째 ids_count'.format(ids_count + 1))

                        # Memo Line
                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:0:memoLineNameId::content"]
                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:1:memoLineNameId::content"]
                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:2:memoLineNameId::content"]

                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:0:descriptionId::content"]
                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:1:descriptionId::content"]
                        # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:2:descriptionId::content"]
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:memoLineNameId::content"]'.format(ids_count)).send_keys(
                            ars.result[issue_cnt].InvoiceDetails[ids_count].MemoLine)
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:memoLineNameId::content"]'.format(ids_count)).send_keys(
                            Keys.ENTER)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Memo Line 입력 완료')

                        # Description
                        # Memo Line보다 먼저 입력하면 나중에 Memo Line 내용으로 바뀜
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:descriptionId::content"]'.format(ids_count)).clear()
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:descriptionId::content"]'.format(ids_count)).send_keys(
                            ars.result[issue_cnt].InvoiceDetails[ids_count].Description)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Description 입력 완료')

                        # Quantity
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:quantity::content"]'.format(ids_count)).send_keys(
                            ars.result[issue_cnt].InvoiceDetails[ids_count].Quantity)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Quantity 완료')

                        # Unit Price
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:sellingPrice::content"]'.format(ids_count)).send_keys(
                            ars.result[issue_cnt].InvoiceDetails[ids_count].UnitPrice)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Unit Price 완료')

                        # Tax Classification : 과세유형
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:taxClassificationCodeId::content"]'.format(ids_count)).send_keys(
                            ars.result[issue_cnt].InvoiceDetails[ids_count].TaxClassification)
                        driver.find_element(By.XPATH,
                                            '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:AT1:_ATp:table1:{}:taxClassificationCodeId::content"]'.format(ids_count)).send_keys(
                            Keys.ENTER)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Tax Classification 완료')

                        ids_count = ids_count + 1



                    # save
                    wait.until(EC.element_to_be_clickable((By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:saveMenu"]/table/tbody/tr/td[1]/a/span')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:TCF:0:ap1:saveMenu"]/table/tbody/tr/td[1]/a/span').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Save 완료')

                    # Action → Edit Distribution
                    time.sleep(10)
                    wait.until(EC.element_to_be_clickable((By.XPATH,
                                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:m1"]/div/table/tbody/tr/td[2]/a').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Action 버튼 클릭 완료')

                    # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:cmi7"]/td[2]').click()
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi7"]/td[2]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi7"]/td[2]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Edit Distribution 클릭 완료')

                    # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:cb7"]').click()
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb7"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb7"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Alert창 Yes 클릭 완료')

                    # Edit Distributions - Distributions - Revenue
                    # # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1CS::content"]').clear()
                    # # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1CS::content"]').send_keys('HX-999-99999-SELHLDM-413131-0000-V170016-000000-00000')
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1CS::content"]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1CS::content"]').clear()
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1CS::content"]').send_keys(
                    #     revenu)
                    # # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Edit Distributions > Revenue 입력 완료')

                    # Edit Distributions - Distributions - 우측 돋보기 버튼 클릭
                    # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1KBIMG::icon"]
                    # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1KBIMG::icon"]
                    # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:3:kf1KBIMG::icon"]
                    # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:5:kf1KBIMG::icon"]
                    # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:7:kf1KBIMG::icon"]
                    # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:9:kf1KBIMG::icon"]

                    ids_count2 = 0
                    while ids_count2 < len(ars.result[issue_cnt].InvoiceDetails):

                        # //*[@id="_afrFilter_FOpt1_afr__FOr1_afr_0_afr__FONSr2_afr_0_afr_MAnt2_afr_2_afr_pt1_afr_Trans1_afr_0_afr_ap110_afr_DistTF1_afr_1_afr_AT2_afr__ATp_afr_table2_afr_c38::content"]
                        # //*[@id="_afrFilter_FOpt1_afr__FOr1_afr_0_afr__FONSr2_afr_0_afr_MAnt2_afr_2_afr_pt1_afr_Trans1_afr_0_afr_ap110_afr_DistTF1_afr_1_afr_AT2_afr__ATp_afr_table2_afr_c40::content"]

                        # /html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[2]/span/input
                        # /html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input

                        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_afrFilter_FOpt1_afr__FOr1_afr_0_afr__FONSr2_afr_0_afr_MAnt2_afr_1_afr_pt1_afr_Trans1_afr_0_afr_ap110_afr_DistTF1_afr_1_afr_AT2_afr__ATp_afr_table2_afr_c38::content"]')))
                        time.sleep(2)

                        # Line Number
                        # //*[@id="_afrFilter_FOpt1_afr__FOr1_afr_0_afr__FONSr2_afr_0_afr_MAnt2_afr_2_afr_pt1_afr_Trans1_afr_0_afr_ap110_afr_DistTF1_afr_1_afr_AT2_afr__ATp_afr_table2_afr_c38::content"]
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[2]/span/input').clear()
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[2]/span/input').send_keys(ids_count2 + 1)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Edit Distributions > Line Number')

                        # Account Class
                        # //*[@id="_afrFilter_FOpt1_afr__FOr1_afr_0_afr__FONSr2_afr_0_afr_MAnt2_afr_2_afr_pt1_afr_Trans1_afr_0_afr_ap110_afr_DistTF1_afr_1_afr_AT2_afr__ATp_afr_table2_afr_c40::content"]
                        # /html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input').clear()
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input').send_keys('Revenue')
                        time.sleep(2)
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[4]/span/input').send_keys(Keys.ENTER)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Edit Distributions > Account Class')

                        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1KBIMG::icon"]'.format(ids_count2 * 2 + 1))))
                        # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1KBIMG::icon"]'.format(ids_count2 * 2 + 1)).click()
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[2]/table/tbody/tr/td[6]/span/span/div/table/tbody/tr/td[3]/a/img')))
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[1]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/div[2]/div/div[2]/table/tbody/tr/td[6]/span/span/div/table/tbody/tr/td[3]/a/img').click()
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Edit Distributions > 돋보기 버튼 클릭 완료')

                        
                        # Edit Distributions - Distributions - RESP Center
                        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value20::content"]')))
                        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1SPOP_query:value20::content"]'.format(ids_count2 * 2 + 1))))
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/span/span/input')))
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/span/span/input').clear()
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/span/span/input').send_keys('99999')
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Distributions > RESP Center 입력 완료')

                        # Edit Distributions - Distributions - RESP Center 클릭
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]')))
                        time.sleep(1)
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]').click()
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Distributions > RESP Center 클릭 완료')

                        # Edit Distributions - Distributions - Cost Center 입력
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[5]/td[2]/table/tbody/tr/td[1]/span/span/input')))
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[5]/td[2]/table/tbody/tr/td[1]/span/span/input').clear()
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[5]/td[2]/table/tbody/tr/td[1]/span/span/input').send_keys(ars.result[issue_cnt].InvoiceDetails[ids_count2].CostCenter)
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Distributions > Cost Center 입력 완료')

                        # Edit Distributions - Distributions - Cost Center 클릭
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]')))
                        time.sleep(1)
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]').click()
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Distributions > Cost Center 클릭 완료')

                        # # Edit Distributions - Distributions - Account 입력
                        # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1SPOP_query:value30::content"]'.format(ids_count2 * 2 + 1)).clear()
                        # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1SPOP_query:value30::content"]'.format(ids_count2 * 2 + 1)).send_keys(
                        #     ars.result[issue_cnt].InvoiceDetails[ids_count2].Account)
                        # time.sleep(2)
                        # logger.LoggerFactory._LOGGER.info('Distributions > Account 입력 완료')

                        # # Edit Distributions - Distributions - Account 클릭
                        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1SPOP_query:value40::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]'.format(ids_count2 * 2 + 1))))
                        # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1SPOP_query:value40::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]'.format(ids_count2 * 2 + 1)).click()
                        # time.sleep(2)
                        # logger.LoggerFactory._LOGGER.info('Distributions > Account 클릭 완료')

                        # Edit Distributions - Distributions - Project 입력
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[8]/td[2]/table/tbody/tr/td[1]/span/span/input')))
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[8]/td[2]/table/tbody/tr/td[1]/span/span/input').clear()
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[8]/td[2]/table/tbody/tr/td[1]/span/span/input').send_keys(ars.result[issue_cnt].InvoiceDetails[ids_count2].Project)  
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Distributions > Project 입력 완료')

                        # Edit Distributions - Distributions - Project 클릭
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]')))
                        time.sleep(1)
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[5]/div/div/table/tbody/tr/td/div/div[1]/ul/li/div/div[2]/div/span[2]').click()
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Distributions > Project 클릭 완료')


                        # Edit Distributions - Distributions - OK 버튼 클릭
                        # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1SEl"]'.format(ids_count2 * 2 + 1))))
                        # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:{}:kf1SEl"]'.format(ids_count2 * 2 + 1)).click()
                        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/span/button[3]')))
                        driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/span/button[3]').click()
                        
                        time.sleep(2)
                        logger.LoggerFactory._LOGGER.info('Distributions > OK 버튼 클릭 완료')

                        ids_count2 = ids_count2 + 1



                    # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1KBIMG::icon"]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1KBIMG::icon"]').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Edit Distributions > 돋보기 버튼 클릭 완료')

                    # # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value20::content"]
                    # # Edit Distributions - Distributions - RESP Center
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value20::content"]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value20::content"]').clear()
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value20::content"]').send_keys(
                    #     '99999')
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Distributions > RESP Center 입력 완료')

                    # # Edit Distributions - Distributions - RESP Center 클릭
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value20::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value20::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Distributions > RESP Center 클릭 완료')

                    # # Edit Distributions - Distributions - Cost Center
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value30::content"]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value30::content"]').clear()
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value30::content"]').send_keys(
                    #     cost_center)  # 전문솔루션그룹
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Distributions > Cost Center 입력 완료')

                    # # Edit Distributions - Distributions - Cost Center 클릭
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value30::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value30::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Distributions > Cost Center 클릭 완료')

                    # # Edit Distributions - Distributions - Project
                    # # /html/body/div[1]/form/div[2]/div[2]/div[3]/div[1]/table/tbody/tr/td/div/div/table/tbody/tr[2]/td[2]/div/div[1]/table/tbody/tr[2]/td[2]/div/div/div[2]/div/table/tbody/tr/td/table/tbody/tr[8]/td[2]/table/tbody/tr/td[1]/span/span/input
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value60::content"]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value60::content"]').clear()
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value60::content"]').send_keys(
                    #     project)  # 대외
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Distributions > Project 입력 완료')

                    # # //*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value60::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value60::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SPOP_query:value60::_fndSuggestPopup_sugg_ListOfValues_0"]/div/div[2]/div/span[2]').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Distributions > Project 클릭 완료')

                    # # Edit Distributions - Distributions - OK 버튼 클릭
                    # wait.until(EC.element_to_be_clickable((By.XPATH,
                    #                                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SEl"]')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:DistTF1:1:AT2:_ATp:table2:1:kf1SEl"]').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('Distributions > OK 버튼 클릭 완료')
                    


                    # Save and Close
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb5"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cb5"]').click()
                    time.sleep(3)  # 없애면 오류발생

                    # Save 우측 화살표버튼 클릭
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:saveMenu::popEl"]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:saveMenu::popEl"]').click()
                    time.sleep(3)  # 없애면 오류발생
                    logger.LoggerFactory._LOGGER.info('Save 우측 화살표버튼 클릭 완료')

                    # Save and Close
                    wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi10"]/td[2]')))
                    driver.find_element(By.XPATH,
                                        '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:pt1:Trans1:0:ap110:cmi10"]/td[2]').click()
                    time.sleep(3)  # 없애면 오류발생
                    logger.LoggerFactory._LOGGER.info('Save and Close 클릭 완료')

                    # Billing 화면에서 Alert창 내 OK버튼 클릭
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]')))
                    driver.find_element(By.XPATH, '//*[@id="_FOd1::msgDlg::cancel"]').click()
                    time.sleep(3)  # 없애면 오류발생
                    logger.LoggerFactory._LOGGER.info('Billing 화면에서 Alert창 내 OK버튼 클릭 완료')

                    ws2['S{}'.format(row_count)] = 'Y'
                    ars.result[issue_cnt].RegisterYN = 'Y'
                    logger.LoggerFactory._LOGGER.info('AR 발행완료 : {}'.format(ars.result[issue_cnt].BillToName))
                    wb.save(destination)

                    # # MBP_Admin 고객사 > 청구내역 관리 내 업데이트 프로세스
                    # logger.LoggerFactory._LOGGER.info('세금계산서 발행 상태 업데이트 : [{}]'.format(BillToName))
                    # update_status_after_writeAR = query_config['update_status_after_writeAR']
                    # cursor.execute(update_status_after_writeAR.format(cust_sn, this_year, this_month))
                    # db_commit(con)
                    # logger.LoggerFactory._LOGGER.info('세금계산서 발행 상태 업데이트 완료')

                except Exception as ex:
                    logger.LoggerFactory._LOGGER.info('error msg : {}'.format(ex))
                    logger.LoggerFactory._LOGGER.info('AR 발행 중 오류 발생 {}회 발생 : {}'.format(re_count, ars.result[issue_cnt].BillToName))
                    # ws2['S{}'.format(row_count)] = 'E'
                    # ars.result[issue_cnt].RegisterYN = 'E'
                    # wb.save(destination)
                    time.sleep(2)

                    re_count = re_count + 1

                    if re_count > 3 :
                        logger.LoggerFactory._LOGGER.info('AR 발행 중 오류 발생 3회 발생으로 에러처리 : {}'.format(ars.result[issue_cnt].BillToName))
                        ws2['S{}'.format(row_count)] = 'E'
                        ars.result[issue_cnt].RegisterYN = 'E'
                        wb.save(destination)

                    driver.refresh()

                    # # 방법1) 기존
                    # # AR Transaction 이동
                    # # 햄버거 버튼 클릭
                    # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_UISmmLink::icon"]')))
                    # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_UISmmLink::icon"]').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('햄버거 버튼 클릭 완료')

                    # # Receivables 메뉴 클릭
                    # wait.until(EC.element_to_be_clickable(
                    #     (By.XPATH, '//*[@id="_FOpt1:_UISnvr:0:nv_itemNode_receivables_billing"]/span')))
                    # driver.find_element(By.XPATH,
                    #                     '//*[@id="_FOpt1:_UISnvr:0:nv_itemNode_receivables_billing"]/span').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('billing 메뉴 클릭 완료')

                    # # changes will be lost. Do you want to continue?
                    # wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAyes"]')))
                    # driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAyes"]').click()
                    # time.sleep(2)
                    # logger.LoggerFactory._LOGGER.info('changes will be lost. Do you want to continue? >>> Yes 클릭 완료')

                    # 방법2) 신규 - 오류내용 삭제
                    logger.LoggerFactory._LOGGER.info("오류 AR 삭제 : [{}]".format(ars.result[issue_cnt].BillToName))

                    # Delete 버튼 클릭
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a')))
                    driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:commandToolbarButton1"]/a').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info('Delete 버튼 클릭 완료')

                    # You're about to delete incomplete transaction. Do you want to continue?
                    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]')))
                    driver.find_element(By.XPATH, '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:Trans1:0:ap110:dialog3::yes"]').click()
                    time.sleep(2)
                    logger.LoggerFactory._LOGGER.info("You're about to delete incomplete transaction. Do you want to continue? >>> Yes 클릭 완료")

                    continue

            else:
                logger.LoggerFactory._LOGGER.info('{}번쨰 : {} 기발행 완료'.format(row_count - ids_count, BillToName))
        
            # billStatement_row = billStatement_row + 1
        
        # ws4 = wb['업체']
        
        # wb.save(destination)
        # logger.LoggerFactory._LOGGER.info('ERP AR_Transaction 발행 종료')



            
            # while ids_count < len(ars.result[issue_cnt].InvoiceDetails):
            #     logger.LoggerFactory._LOGGER.info('{}번째 ids_count'.format(ids_count + 1))

                # ids_count = ids_count + 1

            
            
            # top_left_cell = ws2['S{}'.format(row_count)]
            # top_left_cell.value = 'Y'
            # ws2['S{}'.format(row_count)] = 'N'
            row_count = row_count + len(ars.result[issue_cnt].InvoiceDetails)
            issue_cnt = issue_cnt + 1
        
        wb.save(destination)
        logger.LoggerFactory._LOGGER.info('ERP AR_Transaction 발행 종료')



    else:
        logger.LoggerFactory._LOGGER.info('설치를 통해 사용자 정보 등록 부탁드립니다.')




    # logger.LoggerFactory._LOGGER.info('발행할 AR 갯수 : {}개, {}번째 row, {}번째 배열'.format(len(ars.result), row, count))
    # logger.LoggerFactory._LOGGER.info('발행할 AR 갯수 : {}개, {}번째 row, {}번째 배열'.format(len(ars.result), row, count))


    



    # print(ars.toJsonARs())
    # print(len(ars.result))

    # with open('ars.json', 'w') as fp:
    #     json.dump(ars.toJsonARs(), fp, ensure_ascii=False)
    
    # with open('ars.json', 'r') as f:
    #     after = json.load(f)
    # print(after)


    
    



processing_ERP_AR_Transaction(0, 0, 0, 0)